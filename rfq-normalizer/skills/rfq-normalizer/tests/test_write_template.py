"""Tests for write_template extra-column preservation (v0.7 Fix 4)."""
from __future__ import annotations

from write_template import write, COLUMNS

CANONICAL = [h for h, _ in COLUMNS]


def _read_headers(path):
    from openpyxl import load_workbook
    ws = load_workbook(path).active
    return [c.value for c in ws[1]]


def _read_row(path, row_idx=2):
    from openpyxl import load_workbook
    ws = load_workbook(path).active
    headers = [c.value for c in ws[1]]
    values = [c.value for c in ws[row_idx]]
    return dict(zip(headers, values))


def test_extra_columns_appended_after_canonical(tmp_path):
    out = tmp_path / "o.xlsx"
    rows = [{
        "MPN": "HUA723020ALA640", "Quantity": 1, "Capacity": "2TB",
        "Serial": "JK1101", "Tested": "Yes", "Source": "ListA",
        # internal keys that must NOT become columns:
        "size": "2TB", "drive_type": "HDD",
        "_provenance": {"Capacity": {"source": "regex"}},
    }]
    write(rows, out)
    headers = _read_headers(out)

    assert headers[:len(CANONICAL)] == CANONICAL
    assert headers[len(CANONICAL):] == ["Serial", "Tested", "Source"]
    # internal keys excluded
    assert "size" not in headers and "drive_type" not in headers
    assert "_provenance" not in headers


def test_extra_column_values_intact(tmp_path):
    out = tmp_path / "o.xlsx"
    rows = [{"MPN": "X", "Quantity": 1, "Serial": "SN-9", "Tested": "Pass"}]
    write(rows, out)
    row = _read_row(out)
    assert row["Serial"] == "SN-9"
    assert row["Tested"] == "Pass"


def test_extra_column_order_is_union_first_seen(tmp_path):
    out = tmp_path / "o.xlsx"
    rows = [
        {"MPN": "A", "Quantity": 1, "Serial": "1"},
        {"MPN": "B", "Quantity": 1, "Location": "Bin 3", "Serial": "2"},
    ]
    write(rows, out)
    headers = _read_headers(out)
    assert headers[len(CANONICAL):] == ["Serial", "Location"]


def test_no_extra_columns_keeps_canonical_only(tmp_path):
    out = tmp_path / "o.xlsx"
    rows = [{"MPN": "A", "Quantity": 1, "Capacity": "2TB"}]
    write(rows, out)
    assert _read_headers(out) == CANONICAL
