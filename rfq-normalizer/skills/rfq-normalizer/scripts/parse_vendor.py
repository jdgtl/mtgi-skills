#!/usr/bin/env python3
"""Parse a vendor RFQ spreadsheet (xlsx or csv) and emit normalized rows.

Vendor sheets often carry title/banner rows above the real header and a
TOTAL/summary row at the bottom. This parser auto-detects the header row,
discards banners, and drops aggregate footer rows so they never become bogus
line items.

Output: prints JSON to stdout with shape:
    {
      "headers": ["Part #", "Qty", "Description", ...],
      "rows": [{"Part #": "ABC123", "Qty": 10, ...}, ...],
      "sheet_names": ["Sheet1", ...],          # xlsx only
      "header_row_index": 4,                    # 1-based row the header was found on
      "skipped_banner_rows": ["Evolution E-Cycle", ...],
      "dropped_summary_rows": [{"row": 14, "text": "TOTAL: 1403 drives | 20941"}]
    }

Usage:
    python parse_vendor.py path/to/vendor.xlsx [--sheet "Sheet1"] [--header-row N]
"""
from __future__ import annotations
import argparse
import json
import re
import sys
from pathlib import Path

# Header tokens we expect to see in a real header row. Used as a scoring bonus
# so a banner row ("Combined Drive Inventory") doesn't beat the real header.
KNOWN_HEADER_ALIASES = {
    "part", "part#", "part #", "p/n", "pn", "mpn", "model", "sku",
    "qty", "quantity", "count", "units",
    "price", "bid", "cost", "winning",
    "capacity", "size", "interface", "protocol", "form", "factor",
    "condition", "grade", "health", "tested",
    "serial", "sn", "asset", "brand", "manufacturer", "mfg", "mfr",
    "drive", "type", "outcome", "result", "status",
    "description", "desc", "notes", "source", "location",
}

_TOTAL_RE = re.compile(r"^(grand\s+)?total\b", re.I)
_SUM_RE = re.compile(r"^sum\b", re.I)
_SUMMARY_WORD_RE = re.compile(r"\b(total|subtotal|sum)\b", re.I)
_WORD_RE = re.compile(r"[a-z0-9#/]+")


def _is_blank(v) -> bool:
    return v is None or (isinstance(v, str) and not v.strip())


def _is_number(v) -> bool:
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        try:
            float(v.replace(",", "").replace("$", "").strip())
            return True
        except ValueError:
            return False
    return False


def _score_header_row(cells: list) -> int:
    """Higher = more header-like. Counts short non-numeric string cells, with a
    strong bonus for cells whose words match known header aliases."""
    score = 0
    for cell in cells:
        if _is_blank(cell) or _is_number(cell):
            continue
        text = str(cell).strip()
        if len(text) > 40:
            continue
        score += 1
        words = set(_WORD_RE.findall(text.lower()))
        if words & KNOWN_HEADER_ALIASES:
            score += 3
    return score


def _detect_header_row(rows: list[list], max_scan: int = 15) -> int:
    """Return the 0-based index of the most header-like row in the first
    `max_scan` rows. Falls back to the first non-blank row."""
    best_idx, best_score = None, 0
    for i, row in enumerate(rows[:max_scan]):
        s = _score_header_row(row)
        if s > best_score:
            best_score, best_idx = s, i
    if best_idx is not None:
        return best_idx
    # No row scored — use the first non-blank row.
    for i, row in enumerate(rows):
        if any(not _is_blank(c) for c in row):
            return i
    return 0


def _is_summary_row(values: list) -> bool:
    """Conservatively identify an aggregate/summary row (e.g. a TOTAL footer)."""
    populated = [v for v in values if not _is_blank(v)]
    if not populated:
        return False
    first = str(populated[0]).strip()
    if _TOTAL_RE.match(first) or _SUM_RE.match(first):
        return True
    # Exactly one text cell (containing total/subtotal/sum) plus one numeric cell.
    text_cells = [v for v in populated if not _is_number(v)]
    num_cells = [v for v in populated if _is_number(v)]
    if len(text_cells) == 1 and len(num_cells) == 1 and _SUMMARY_WORD_RE.search(str(text_cells[0])):
        return True
    return False


def _extract(raw_rows: list[list], header_row: int | None) -> dict:
    """Shared extraction for xlsx and csv once rows are materialized as lists.

    `header_row` is a 1-based override; None means auto-detect.
    """
    if not raw_rows:
        return {
            "headers": [], "rows": [], "header_row_index": 0,
            "skipped_banner_rows": [], "dropped_summary_rows": [],
        }

    if header_row is not None:
        header_idx = max(0, min(header_row - 1, len(raw_rows) - 1))
    else:
        header_idx = _detect_header_row(raw_rows)

    header_cells = raw_rows[header_idx]
    # Map by original column index so gaps in the header don't misalign data.
    header_map = [
        (i, str(h).strip())
        for i, h in enumerate(header_cells)
        if not _is_blank(h)
    ]
    headers = [h for _, h in header_map]

    banners = [
        " | ".join(str(c).strip() for c in row if not _is_blank(c))
        for row in raw_rows[:header_idx]
        if any(not _is_blank(c) for c in row)
    ]

    rows: list[dict] = []
    dropped: list[dict] = []
    for offset, raw in enumerate(raw_rows[header_idx + 1:]):
        if not any(not _is_blank(c) for c in raw):
            continue  # blank row
        if _is_summary_row(raw):
            dropped.append({
                "row": header_idx + 2 + offset,  # 1-based original row number
                "text": " | ".join(str(c).strip() for c in raw if not _is_blank(c)),
            })
            continue
        rows.append({h: (raw[i] if i < len(raw) else None) for i, h in header_map})

    return {
        "headers": headers,
        "rows": rows,
        "header_row_index": header_idx + 1,  # 1-based for reporting
        "skipped_banner_rows": banners,
        "dropped_summary_rows": dropped,
    }


def parse_xlsx(path: Path, sheet: str | None, header_row: int | None = None) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    target = wb[sheet] if sheet else wb[sheet_names[0]]

    raw_rows = [list(r) for r in target.iter_rows(values_only=True)]
    result = _extract(raw_rows, header_row)
    result["sheet_names"] = sheet_names
    return result


def parse_csv(path: Path, header_row: int | None = None) -> dict:
    import csv

    with path.open("r", newline="", encoding="utf-8-sig") as f:
        raw_rows = [list(r) for r in csv.reader(f)]
    result = _extract(raw_rows, header_row)
    result["sheet_names"] = []
    return result


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("path", type=Path)
    ap.add_argument("--sheet", default=None)
    ap.add_argument(
        "--header-row", type=int, default=None,
        help="1-based row number of the header (overrides auto-detection)",
    )
    args = ap.parse_args()

    if not args.path.exists():
        print(f"ERROR: file not found: {args.path}", file=sys.stderr)
        return 1

    suffix = args.path.suffix.lower()
    if suffix == ".xlsx":
        out = parse_xlsx(args.path, args.sheet, header_row=args.header_row)
    elif suffix == ".csv":
        out = parse_csv(args.path, header_row=args.header_row)
    else:
        print(f"ERROR: unsupported file type {suffix}", file=sys.stderr)
        return 1

    json.dump(out, sys.stdout, default=str, indent=2)
    return 0


if __name__ == "__main__":
    sys.exit(main())
