#!/usr/bin/env python3
"""Write the final normalized xlsx in MTGI template format + a provenance JSON.

Input (stdin):
    {
      "rows": [
        {
          "MPN": "...", "Quantity": 12, "Bid Price (USD)": 412.50,
          "Condition": "new", "Description": "...", "Outcome": "won",
          "Outcome Date": "2026-04-08", "Winning Bid (USD)": 412.50,
          "Capacity": "1.6TB", "Interface": "SATA", "Drive Type": "SSD",
          "Form Factor": "2.5in", "Manufacturer": "Intel",
          "_provenance": {"Capacity": {"source": "regex", "confidence": 0.95}, ...}
        }
      ],
      "output_path": "vendor-normalized.xlsx"
    }

Any additional keys on the rows (e.g. Serial, Tested, Source) are preserved as
extra columns after the 13 canonical ones, so the MTGI wizard captures them as
custom_fields rather than the skill dropping them. Internal keys (_provenance
and the internal spec keys) are never written as columns.

Output:
    <output_path>           — xlsx in MTGI template format
    <output_path>.provenance.json  — per-cell provenance log
"""
from __future__ import annotations
import argparse
import json
import sys
from pathlib import Path

COLUMNS = [
    ("MPN",                True),
    ("Quantity",           True),
    ("Bid Price (USD)",    True),
    ("Condition",          False),
    ("Description",        False),
    ("Outcome",            False),
    ("Outcome Date",       False),
    ("Winning Bid (USD)",  False),
    ("Capacity",           False),
    ("Interface",          False),
    ("Drive Type",         False),
    ("Form Factor",        False),
    ("Manufacturer",       False),
]

REQUIRED_FILL = {"argb": "FF0D9488", "font": "FFFFFFFF"}
OPTIONAL_FILL = {"argb": "FFE5E7EB", "font": "FF1F2937"}

# Per-row keys that are internal plumbing, not output columns. These never
# become extra columns: the internal spec keys the canonicalizer maps into the
# canonical columns. Plus, ANY underscore-prefixed key is stripped from the
# delivered xlsx (e.g. _provenance and the engine's _mpn/_source/_confidence/
# _flags helpers) — those are retained only in the provenance + needs-review
# sidecars (v0.9 Q3: deliverables stay clean).
INTERNAL_KEYS = {
    "size", "interface", "drive_type", "form_factor", "manufacturer",
}

# Default width for preserved extra/custom columns.
EXTRA_COL_WIDTH = 16


def _extra_headers(rows: list[dict]) -> list[str]:
    """Union of non-canonical, non-internal row keys in first-seen order.

    These are preserved (e.g. Serial, Tested, Source) so the MTGI wizard can
    capture them as custom_fields instead of the skill silently dropping them.
    """
    canonical = {h for h, _ in COLUMNS}
    extras: list[str] = []
    for row in rows:
        for key in row:
            if key.startswith("_") or key in canonical or key in INTERNAL_KEYS or key in extras:
                continue
            extras.append(key)
    return extras


def write(rows: list[dict], output_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.freeze_panes = "A2"

    # 13 canonical columns + any preserved extras (styled as optional).
    extra_headers = _extra_headers(rows)
    all_columns = list(COLUMNS) + [(h, False) for h in extra_headers]

    # Header row
    for i, (header, required) in enumerate(all_columns, start=1):
        cell = ws.cell(row=1, column=i, value=header)
        fill = REQUIRED_FILL if required else OPTIONAL_FILL
        cell.font = Font(name="Calibri", size=11, bold=required, color=fill["font"])
        cell.fill = PatternFill("solid", fgColor=fill["argb"])
        cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        cell.border = Border(bottom=Side(style="medium", color="FF0F766E"))

    # Column widths — canonical widths, then a default for each extra column.
    widths = [22, 12, 16, 18, 36, 22, 14, 18, 12, 18, 14, 14, 16]
    widths += [EXTRA_COL_WIDTH] * len(extra_headers)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.row_dimensions[1].height = 24

    # Data rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, (header, _) in enumerate(all_columns, start=1):
            value = row.get(header)
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = Font(name="Consolas" if header == "MPN" else "Calibri", size=11)
            cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
            if header in ("Bid Price (USD)", "Winning Bid (USD)"):
                cell.number_format = '"$"#,##0.00;[Red]-"$"#,##0.00;""'
                cell.alignment = Alignment(vertical="center", horizontal="right")
            elif header == "Quantity":
                cell.number_format = "#,##0"
                cell.alignment = Alignment(vertical="center", horizontal="right")

    wb.save(output_path)


# ── Needs-review report (v0.8 Change 5) ──────────────────────────────────────

# Core columns enrichment targets. The first three apply to every part; the
# storage specs only apply to storage parts (blanked-by-gate rows are skipped).
ALWAYS_CORE = ["MPN", "Manufacturer", "Condition"]
STORAGE_CORE = ["Capacity", "Interface", "Drive Type", "Form Factor"]
CORE_COLUMNS = ALWAYS_CORE + STORAGE_CORE


def _is_empty(v) -> bool:
    return v is None or (isinstance(v, str) and not v.strip())


def needs_review(rows: list[dict]) -> list[dict]:
    """Return one entry per row that still has a blank core column, a
    low-confidence core value, or an unresolved MPN. Storage specs that were
    intentionally blanked for a non-storage part (provenance note) don't count.
    """
    out: list[dict] = []
    for i, row in enumerate(rows, start=1):
        prov = row.get("_provenance") or {}
        missing: list[str] = []
        low_conf: list[str] = []
        for col in CORE_COLUMNS:
            col_prov = prov.get(col) or {}
            note = str(col_prov.get("note") or "").lower()
            if col in STORAGE_CORE and "non-storage" in note:
                continue  # intentionally blank — not a gap
            if _is_empty(row.get(col)):
                missing.append(col)
            elif col_prov.get("tagged_low_confidence"):
                low_conf.append(col)
        candidate = row.get("_candidate_real_mpn")
        unresolved_mpn = bool(candidate) or bool(row.get("_mpn_unresolved"))
        flags = str(row.get("_flags") or "")
        if missing or low_conf or unresolved_mpn or flags:
            out.append({
                "row": i,
                "MPN": row.get("MPN"),
                "Manufacturer": row.get("Manufacturer"),
                "missing_fields": missing,
                "low_confidence_fields": low_conf,
                "candidate_real_mpn": candidate,
                "unresolved_mpn": unresolved_mpn,
                "_confidence": row.get("_confidence") or "",
                "_flags": flags,
            })
    return out


def review_summary(rows: list[dict], review: list[dict]) -> str:
    blank = sum(1 for r in review if r["missing_fields"] or r["low_confidence_fields"])
    unresolved = sum(1 for r in review if r["unresolved_mpn"])
    return (f"{len(review)} of {len(rows)} rows need review "
            f"({blank} with blank/low-confidence specs, {unresolved} unresolved MPNs).")


def write_needs_review_csv(review: list[dict], path: Path) -> None:
    import csv
    with path.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["row", "MPN", "Manufacturer", "missing_fields",
                    "low_confidence_fields", "_confidence", "_flags",
                    "candidate_real_mpn"])
        for r in review:
            w.writerow([
                r["row"], r["MPN"] or "", r["Manufacturer"] or "",
                "; ".join(r["missing_fields"]),
                "; ".join(r["low_confidence_fields"]),
                r.get("_confidence", ""), r.get("_flags", ""),
                r["candidate_real_mpn"] or "",
            ])


def _needs_review_path(output_path: Path) -> Path:
    stem = output_path.stem
    base = stem[:-len("-normalized")] if stem.endswith("-normalized") else stem
    return output_path.with_name(f"{base}-needs-review.csv")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default=None, help="JSON file (default: stdin)")
    args = ap.parse_args()

    raw = json.load(open(args.input) if args.input else sys.stdin)
    rows = raw["rows"]
    output_path = Path(raw["output_path"])
    output_path.parent.mkdir(parents=True, exist_ok=True)

    write(rows, output_path)

    # Provenance log
    provenance = {
        i + 1: row.get("_provenance", {}) for i, row in enumerate(rows)
    }
    prov_path = output_path.with_suffix(output_path.suffix + ".provenance.json")
    with prov_path.open("w") as f:
        json.dump(provenance, f, indent=2, default=str)

    # Needs-review report — rows with blank/low-confidence core columns or an
    # unresolved MPN. Always written (header-only when nothing needs review).
    review = needs_review(rows)
    review_path = _needs_review_path(output_path)
    write_needs_review_csv(review, review_path)
    summary = review_summary(rows, review)

    print(json.dumps({
        "xlsx": str(output_path),
        "provenance": str(prov_path),
        "needs_review": str(review_path),
        "needs_review_count": len(review),
        "summary": summary,
    }))
    return 0


if __name__ == "__main__":
    sys.exit(main())
